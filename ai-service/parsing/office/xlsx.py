"""XLSX 解析器 — 增强版：每个 Sheet 独立为 TableBlock。

相比旧版改进：
    - 表格不再序列化为 Markdown 字符串混入文本流
    - 每个 Sheet 独立为 TableBlock
    - 额外生成 Sheet 摘要 TextBlock
"""

from openpyxl import load_workbook

from common import get_logger
from common.utils import estimate_tokens
from parsing.base import BaseParser
from parsing.models import ParsedDocument, DocumentMetadata, TextBlock, TableBlock

logger = get_logger(__name__)

MAX_ROWS = 500  # 每个 Sheet 最大读取行数


class XlsxParser(BaseParser):
    """XLSX 文件解析器。将每个非空 Sheet 转换为独立的 TableBlock。"""

    async def _do_parse(self, file_path: str) -> ParsedDocument:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        blocks: list[TextBlock | TableBlock] = []
        total_sheets = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(max_row=MAX_ROWS, values_only=True))
            if not rows:
                continue

            # 清洗空行/空列
            clean_rows = [[str(c) if c is not None else "" for c in row] for row in rows]
            clean_rows = [r for r in clean_rows if any(c.strip() for c in r)]
            if not clean_rows:
                continue

            total_sheets += 1

            # Sheet 摘要文本块
            token_est = estimate_tokens("\n".join(" ".join(r) for r in clean_rows))
            blocks.append(TextBlock(
                text=f"Sheet: {sheet_name} — {len(clean_rows)} 行数据",
                page_num=total_sheets - 1,
                layout_type="title",
                level=2,
                metadata={"sheet_name": sheet_name},
            ))

            # 表格块
            html = _to_html_table(clean_rows)
            description = _describe_sheet(sheet_name, clean_rows, token_est)
            blocks.append(TableBlock(
                html=html,
                description=description,
                page_num=total_sheets - 1,
                caption=None,
                metadata={
                    "sheet_name": sheet_name,
                    "row_count": len(clean_rows),
                    "col_count": len(clean_rows[0]) if clean_rows else 0,
                },
            ))

        logger.info(f"XLSX 解析完成: {file_path}, sheets={total_sheets}, blocks={len(blocks)}")
        return ParsedDocument(
            blocks=blocks,
            metadata=DocumentMetadata(
                file_type="xlsx",
                file_name=file_path,
                page_count=total_sheets,
            ),
        )

    def supports(self, file_type: str) -> bool:
        return file_type == "xlsx"


def _to_html_table(data: list[list[str]]) -> str:
    """二维数组 → HTML <table>。"""
    if not data:
        return ""
    rows = ["<table>"]
    # 表头
    rows.append("<thead><tr>")
    for cell in data[0]:
        rows.append(f"<th>{cell}</th>")
    rows.append("</tr></thead>")
    if len(data) > 1:
        rows.append("<tbody>")
        for row in data[1:]:
            rows.append("<tr>")
            for cell in row:
                rows.append(f"<td>{cell}</td>")
            rows.append("</tr>")
        rows.append("</tbody>")
    rows.append("</table>")
    return "\n".join(rows)


def _describe_sheet(sheet_name: str, rows: list[list[str]], token_est: int) -> str:
    """生成 Sheet 语义描述。"""
    headers = rows[0] if rows else []
    row_count = len(rows) - 1  # 减去表头
    col_count = len(headers)
    return (
        f"Sheet '{sheet_name}'：{row_count} 行 × {col_count} 列"
        f"，表头为 [{', '.join(headers[:8])}]"
        f"，约 {token_est} tokens"
    )
