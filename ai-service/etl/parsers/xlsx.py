"""XLSX 解析器 — 使用 openpyxl 提取单元格数据。"""

from openpyxl import load_workbook

from common import get_logger
from models.document import ParseResult
from etl.common.table_utils import rows_to_markdown
from .base import BaseParser

logger = get_logger(__name__)


class XlsxParser(BaseParser):
    """XLSX 文件解析器。将每个 Sheet 渲染为 Markdown Table 文本。"""

    async def _do_parse(self, file_path: str) -> ParseResult:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        parts: list[str] = []
        tables: list[dict] = []
        total_sheets = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            clean_rows = [[str(c) if c is not None else "" for c in row] for row in rows[:500]]
            clean_rows = [r for r in clean_rows if any(c.strip() for c in r)]
            if not clean_rows:
                continue

            total_sheets += 1
            tables.append({
                "sheet": sheet_name, "headers": clean_rows[0],
                "rows": clean_rows[1:], "row_count": len(clean_rows),
            })
            parts.append(f"## {sheet_name}\n\n{rows_to_markdown(clean_rows)}")

        raw_text = "\n\n".join(parts)
        logger.info(f"XLSX 解析完成: {file_path}, sheets={total_sheets}, chars={len(raw_text)}")
        return ParseResult(
            file_type="xlsx", raw_text=raw_text, page_count=total_sheets,
            tables=tables, metadata={"sheet_count": total_sheets},
        )

    def supports(self, file_type: str) -> bool:
        return file_type == "xlsx"
